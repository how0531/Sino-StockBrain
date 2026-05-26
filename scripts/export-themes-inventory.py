# -*- coding: utf-8 -*-
"""Export every theme (族群/關係鍊) in the brain to an Excel for manual review."""
import os, re, glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

THEMES = r"E:\SinoBrain-data\themes"
OUT_DIR = r"E:\SinoBrain-data\ops"
OUT = os.path.join(OUT_DIR, "themes-inventory.xlsx")

member_re = re.compile(r"^- \[\[tickers/([^\]|]+)\]\]\s*(?:\(([^)]*)\))?", re.M)

rows = []
for path in glob.glob(os.path.join(THEMES, "*.md")):
    raw = open(path, encoding="utf-8").read()
    fm = {}
    m = re.match(r"^---\n(.*?)\n---\n", raw, re.S)
    if m:
        for line in m.group(1).splitlines():
            if ":" in line:
                k, v = line.split(":", 1)
                fm[k.strip()] = v.strip().strip('"')
    members = member_re.findall(raw)  # list of (code, name)
    slug = fm.get("slug", os.path.basename(path)[:-3])
    rows.append({
        "slug": slug,
        "title": fm.get("title", ""),
        "source": fm.get("source", "(seed/unknown)"),
        "relation_kind": fm.get("relation_kind", ""),
        "count": len(members),
        "members": members,
    })

# hand-curated / supply-chain first, then by size desc
def sortkey(r):
    hand = 0 if r["source"] == "hand-curated" else (1 if r["source"] != "cmoney-statementdog" else 2)
    return (hand, -r["count"], r["slug"])
rows.sort(key=sortkey)

wb = openpyxl.Workbook()
hdrfont = Font(bold=True, color="FFFFFF")
hdrfill = PatternFill("solid", fgColor="2F4F6F")
handfill = PatternFill("solid", fgColor="FFF3CD")  # highlight hand-curated

# Sheet 1: 族群清單
ws = wb.active
ws.title = "族群清單"
cols = ["slug", "title", "來源", "relation_kind", "成員數", "成員清單(代號 名稱)"]
ws.append(cols)
for c in range(1, len(cols)+1):
    ws.cell(1, c).font = hdrfont; ws.cell(1, c).fill = hdrfill
for r in rows:
    mtxt = "、".join(f"{c}{(' '+n) if n else ''}" for c, n in r["members"])
    ws.append([r["slug"], r["title"], r["source"], r["relation_kind"], r["count"], mtxt])
    if r["source"] == "hand-curated":
        for c in range(1, len(cols)+1):
            ws.cell(ws.max_row, c).fill = handfill
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 120
ws.freeze_panes = "A2"

# Sheet 2: 族群×成員 (long form for pivot/filter)
ws2 = wb.create_sheet("族群x成員")
cols2 = ["族群slug", "來源", "relation_kind", "成員代號", "成員名稱"]
ws2.append(cols2)
for c in range(1, len(cols2)+1):
    ws2.cell(1, c).font = hdrfont; ws2.cell(1, c).fill = hdrfill
for r in rows:
    for code, name in r["members"]:
        ws2.append([r["slug"], r["source"], r["relation_kind"], code, name])
for w, col in zip([26, 20, 20, 12, 16], "ABCDE"):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"

os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT)

# summary to stdout
from collections import Counter
src = Counter(r["source"] for r in rows)
print(f"themes={len(rows)}  rows(族群x成員)={ws2.max_row-1}")
print("by source:", dict(src))
print("hand-curated:", [r["slug"] for r in rows if r["source"] == "hand-curated"])
print("saved:", OUT)
